import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.Workbook()

header_font = Font(bold=True, color="FFFFFF", size=11, name="Arial")
header_fill = PatternFill("solid", fgColor="1e3a8a")
normal_font = Font(size=10, name="Arial")
thin_border = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
alt_fill = PatternFill("solid", fgColor="f1f5f9")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

def style_row(ws, row, cols, is_alt=False):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = left_wrap if c <= 2 else center
        if is_alt:
            cell.fill = alt_fill

def add_title(ws, row, title, cols):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    cell = ws.cell(row=row, column=1)
    cell.value = title
    cell.font = Font(bold=True, size=12, name="Arial", color="1e3a8a")
    cell.fill = PatternFill("solid", fgColor="eff6ff")
    cell.alignment = Alignment(horizontal="left", vertical="center")
    for c in range(1, cols+1):
        ws.cell(row=row, column=c).border = thin_border

# Sheet 1: Cloud
ws1 = wb.active
ws1.title = "Cloud Deployment"
ws1.sheet_properties.tabColor = "0284c7"
r = 1
add_title(ws1, r, "Cloud Deployment - Compute Requirements", 4); r += 1
for c, h in enumerate(["Component", "Small (up to 100)", "Medium (100-500)", "Large (500+)"], 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 4); r += 1
for i, d in enumerate([
    ['Application Server (vCPU)', '2 vCPU', '4 vCPU', '8 vCPU'],
    ['Application Server (RAM)', '4 GB', '8 GB', '16 GB'],
    ['Database Server (vCPU)', '2 vCPU', '4 vCPU', '8 vCPU'],
    ['Database Server (RAM)', '4 GB', '8 GB', '32 GB'],
    ['Storage (SSD)', '50 GB', '100 GB', '500 GB'],
    ['CDN / Static Assets', 'Included', 'Included', 'Included'],
    ['SSL Certificate', "Let's Encrypt", 'Managed SSL', 'Wildcard SSL'],
    ['Backup Storage', '25 GB', '50 GB', '200 GB'],
]):
    for c, v in enumerate(d, 1): ws1.cell(row=r, column=c, value=v)
    style_row(ws1, r, 4, i%2==0); r += 1

r += 1; add_title(ws1, r, "Recommended Cloud Platforms", 4); r += 1
for c, h in enumerate(["Platform", "Services", "Min Tier", ""], 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 3); r += 1
for i, d in enumerate([
    ['AWS', 'EC2 / RDS / S3 / CloudFront', 't3.medium+'],
    ['Azure', 'App Service / Azure SQL / Blob / CDN', 'B2+'],
    ['Google Cloud', 'Cloud Run / Cloud SQL / GCS / CDN', 'e2-medium+'],
]):
    for c, v in enumerate(d, 1): ws1.cell(row=r, column=c, value=v)
    style_row(ws1, r, 3, i%2==0); r += 1

r += 1; add_title(ws1, r, "Existing Vendor Options", 4); r += 1
for c, h in enumerate(["Vendor", "Services", "", ""], 1):
    ws1.cell(row=r, column=c, value=h)
style_header(ws1, r, 2); r += 1
for i, d in enumerate([
    ['DigitalOcean', 'Droplets (4 GB+) / Managed DB / Spaces'],
    ['Oracle Cloud (OCI)', 'Compute / Autonomous DB / Object Storage'],
    ['IBM Cloud', 'Virtual Servers / Db2 / Cloud Object Storage'],
    ['Linode (Akamai)', 'Dedicated CPU / Managed DB / Object Storage'],
    ['Hetzner', 'Cloud Servers / Managed DB / Storage Box'],
    ["Client's Own Data Center", 'On-premise VM / Bare metal'],
]):
    for c, v in enumerate(d, 1): ws1.cell(row=r, column=c, value=v)
    style_row(ws1, r, 2, i%2==0); r += 1

r += 1; add_title(ws1, r, "Cloud Security Checklist", 4); r += 1
for i, item in enumerate([
    'HTTPS enforced on all endpoints with TLS 1.3',
    'Database accessible only via private VPC/subnet',
    'Environment variables for all secrets (never hardcoded)',
    'Automated daily backups with 30-day retention',
    'Web Application Firewall (WAF) enabled',
    'DDoS protection (AWS Shield / Azure DDoS / Cloud Armor)',
    'Role-based IAM policies for infrastructure access',
    'Container image scanning (if using Docker)',
]):
    ws1.cell(row=r, column=1, value="✓"); ws1.cell(row=r, column=2, value=item)
    ws1.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
    style_row(ws1, r, 4, i%2==0); r += 1

ws1.column_dimensions['A'].width = 30; ws1.column_dimensions['B'].width = 30
ws1.column_dimensions['C'].width = 25; ws1.column_dimensions['D'].width = 25

# Sheet 2: On-Premise
ws2 = wb.create_sheet("On-Premise Server")
ws2.sheet_properties.tabColor = "16a34a"
r = 1
add_title(ws2, r, "On-Premise Hardware Requirements", 3); r += 1
for c, h in enumerate(["Component", "Minimum", "Recommended"], 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 3); r += 1
for i, d in enumerate([
    ['Processor', 'Intel Xeon E-2236 (6C/12T)', 'Intel Xeon Silver 4214 (12C/24T)'],
    ['RAM', '16 GB DDR4 ECC', '32 GB DDR4 ECC'],
    ['Primary Storage', '256 GB NVMe SSD', '512 GB NVMe SSD (RAID 1)'],
    ['Data Storage', '500 GB SAS HDD', '1 TB SAS SSD (RAID 10)'],
    ['Network', '1 Gbps NIC', 'Dual 1 Gbps NIC (bonded)'],
    ['Power Supply', '500W', '800W Redundant PSU'],
    ['UPS Backup', '1 kVA (15 min)', '3 kVA (30 min)'],
    ['Rack Space', '1U', '2U'],
]):
    for c, v in enumerate(d, 1): ws2.cell(row=r, column=c, value=v)
    style_row(ws2, r, 3, i%2==0); r += 1

r += 1; add_title(ws2, r, "Software Stack", 3); r += 1
for c, h in enumerate(["Layer", "Technology", "Version"], 1):
    ws2.cell(row=r, column=c, value=h)
style_header(ws2, r, 3); r += 1
for i, d in enumerate([
    ['Operating System', 'Ubuntu Server LTS / RHEL', '22.04 / 9.x'],
    ['Runtime', 'Node.js (via Deno or Bun)', '20 LTS+'],
    ['Web Server / Proxy', 'Nginx / Caddy', '1.24+ / 2.x'],
    ['Database', 'PostgreSQL', '15+'],
    ['Cache Layer', 'Redis (optional)', '7.x'],
    ['Containerization', 'Docker + Docker Compose', '24.x+'],
    ['Process Manager', 'PM2 / systemd', 'Latest'],
    ['Monitoring', 'Prometheus + Grafana', 'Latest'],
]):
    for c, v in enumerate(d, 1): ws2.cell(row=r, column=c, value=v)
    style_row(ws2, r, 3, i%2==0); r += 1

r += 1; add_title(ws2, r, "On-Premise Security Requirements", 3); r += 1
for i, item in enumerate([
    'Firewall: Allow only ports 80, 443, and SSH (custom port)',
    'Database port (5432) restricted to application server only',
    'OS-level hardening (disable root SSH, key-based auth only)',
    'Automated OS patch management (unattended-upgrades)',
    'RAID configuration for data redundancy',
    'Scheduled automated backups to off-site/NAS storage',
    'Antivirus/malware scanning on server',
    'Physical rack access restricted with biometric/card lock',
]):
    ws2.cell(row=r, column=1, value="✓"); ws2.cell(row=r, column=2, value=item)
    ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
    style_row(ws2, r, 3, i%2==0); r += 1

ws2.column_dimensions['A'].width = 30; ws2.column_dimensions['B'].width = 40; ws2.column_dimensions['C'].width = 35

# Sheet 3: Network & Devices
ws3 = wb.create_sheet("Network & Devices")
ws3.sheet_properties.tabColor = "7c3aed"
r = 1
add_title(ws3, r, "Network Infrastructure", 3); r += 1
for c, h in enumerate(["Requirement", "Specification", ""], 1):
    ws3.cell(row=r, column=c, value=h)
style_header(ws3, r, 2); r += 1
for i, d in enumerate([
    ['Internet Bandwidth (Cloud)', 'Minimum 10 Mbps dedicated uplink'],
    ['LAN Speed (On-Premise)', '100 Mbps minimum, 1 Gbps recommended'],
    ['DNS', 'Custom domain with A/CNAME record configured'],
    ['Static IP', 'Required for on-premise; elastic IP for cloud'],
    ['VPN (optional)', 'Site-to-site VPN for multi-location setups'],
    ['Wi-Fi at Gates', '2.4/5 GHz for QR scanner tablets/kiosks'],
]):
    for c, v in enumerate(d, 1): ws3.cell(row=r, column=c, value=v)
    style_row(ws3, r, 2, i%2==0); r += 1

r += 1; add_title(ws3, r, "Client Device Requirements", 3); r += 1
for c, h in enumerate(["Device", "Purpose", "Min Specification"], 1):
    ws3.cell(row=r, column=c, value=h)
style_header(ws3, r, 3); r += 1
for i, d in enumerate([
    ['Desktop / Laptop', 'Admin Dashboard, Reports', 'Modern browser (Chrome 90+, Edge, Firefox)'],
    ['Tablet (10")', 'Gate check-in kiosk, Self-service', 'Android 10+ / iPad OS 14+ with camera'],
    ['Smartphone', 'Guard mobile app, QR scanning', 'Android 9+ / iOS 14+ with camera'],
    ['Badge Printer', 'Visitor badge printing', 'Thermal printer (100x150mm) or A4 laser'],
    ['Barcode/QR Scanner', 'Rapid check-in (optional)', 'USB or Bluetooth 2D scanner'],
    ['ANPR Camera', 'Vehicle number plate capture', 'IP camera with ANPR/LPR (2MP+, IR)'],
    ['Boom Barrier', 'Automated gate open/close', 'ANPR-integrated boom barrier'],
    ['RFID Reader', 'Recurring vehicle ID', 'UHF RFID reader (865-868 MHz)'],
]):
    for c, v in enumerate(d, 1): ws3.cell(row=r, column=c, value=v)
    style_row(ws3, r, 3, i%2==0); r += 1

ws3.column_dimensions['A'].width = 28; ws3.column_dimensions['B'].width = 35; ws3.column_dimensions['C'].width = 45

# Sheet 4: Data Flow
ws4 = wb.create_sheet("Data Flow & Architecture")
ws4.sheet_properties.tabColor = "dc2626"
r = 1
add_title(ws4, r, "Data Flow Summary", 4); r += 1
for c, h in enumerate(["Source", "Destination", "Protocol / Port", "Purpose"], 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, 4); r += 1
for i, d in enumerate([
    ['Browser / Tablet', 'App Server', 'HTTPS / 443', 'Web app access & API calls'],
    ['App Server', 'PostgreSQL', 'TCP / 5432', 'Database queries & writes'],
    ['App Server', 'Redis', 'TCP / 6379', 'Session cache (optional)'],
    ['App Server', 'Twilio API', 'HTTPS / 443', 'SMS & WhatsApp notifications'],
    ['App Server', 'Resend API', 'HTTPS / 443', 'Email notifications & badges'],
    ['ANPR Camera', 'NVR / Server', 'RTSP / IP', 'Vehicle plate capture feed'],
    ['RFID Reader', 'Gate Controller', 'UHF (865-868 MHz)', 'Vehicle tag identification'],
    ['Gate Tablet', 'Access Point', 'Wi-Fi', 'Check-in/out operations'],
]):
    for c, v in enumerate(d, 1): ws4.cell(row=r, column=c, value=v)
    style_row(ws4, r, 4, i%2==0); r += 1

r += 1; add_title(ws4, r, "Gate Hardware Connectivity", 4); r += 1
for c, h in enumerate(["Device", "Protocol", "Connects To", ""], 1):
    ws4.cell(row=r, column=c, value=h)
style_header(ws4, r, 3); r += 1
for i, d in enumerate([
    ['Tablet / Kiosk', 'Wi-Fi (2.4/5 GHz)', 'Access Point → LAN'],
    ['Badge Printer', 'USB / LAN', 'Gate PC / Tablet'],
    ['QR / Barcode Scanner', 'USB / Bluetooth', 'Gate PC / Tablet'],
    ['ANPR Camera', 'IP / PoE (Ethernet)', 'NVR → Server'],
    ['Boom Barrier', 'Controller (RS-485)', 'ANPR System'],
    ['RFID Reader (UHF)', '865-868 MHz', 'Gate Controller → Server'],
]):
    for c, v in enumerate(d, 1): ws4.cell(row=r, column=c, value=v)
    style_row(ws4, r, 3, i%2==0); r += 1

ws4.column_dimensions['A'].width = 25; ws4.column_dimensions['B'].width = 25
ws4.column_dimensions['C'].width = 30; ws4.column_dimensions['D'].width = 30

# Sheet 5: Manpower
ws5 = wb.create_sheet("Manpower Requirement")
ws5.sheet_properties.tabColor = "f59e0b"
r = 1
add_title(ws5, r, "Manpower Requirement (Client Side)", 3); r += 1
for c, h in enumerate(["Role", "Responsibility", "Required Per Site"], 1):
    ws5.cell(row=r, column=c, value=h)
style_header(ws5, r, 3); r += 1
for i, d in enumerate([
    ['Process Owner', 'Overall ownership of VMS adoption, SOP alignment & compliance', '1'],
    ['Decision Maker / SPOC', 'Approve configurations, escalations & change requests', '1'],
    ['Application Tester', 'Test features, validate workflows, report bugs & feedback', '1-2'],
    ['IT Coordinator', 'Manage network, devices, server access & tech coordination', '1'],
    ['Gate Operator / Security', 'Day-to-day check-in/out, badge printing, QR scanning', 'As per gates'],
]):
    for c, v in enumerate(d, 1): ws5.cell(row=r, column=c, value=v)
    style_row(ws5, r, 3, i%2==0); r += 1

ws5.column_dimensions['A'].width = 25; ws5.column_dimensions['B'].width = 65; ws5.column_dimensions['C'].width = 20

for ws in [ws1, ws2, ws3, ws4, ws5]:
    ws.oddFooter.center.text = "© 2026 Sharvi Infotech | info@sharviinfotech.com | +91 88976 46530"

wb.save("/mnt/documents/VisiGuard-Resource-Requirements.xlsx")
print("Done!")
